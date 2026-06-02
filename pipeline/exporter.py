"""
exporter.py
Exports COBie sheets to a properly formatted .xlsx file.
Follows COBie UK 2012 colour coding:
- Yellow  = required fields
- Orange  = reference fields
- Purple  = internal fields (auto-generated)
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# COBie UK 2012 colour scheme
COLOURS = {
    "header_required":  "FFD700",  # Yellow — required fields
    "header_reference": "FFA500",  # Orange — reference fields
    "header_internal":  "9B59B6",  # Purple — auto fields
    "header_default":   "4A90D9",  # Blue   — standard
    "row_alt":          "F5F5F5",  # Light grey alternating rows
    "critical":         "FFCCCC",  # Red tint for TBC values
    "sheet_tab":        "1F3864",  # Dark blue sheet tabs
}

# Required fields per sheet (yellow header)
REQUIRED_FIELDS = {
    "Contact":   ["Name", "CreatedBy", "CreatedOn", "Email", "Company"],
    "Facility":  ["Name", "CreatedBy", "CreatedOn", "ProjectName", "SiteName"],
    "Floor":     ["Name", "CreatedBy", "CreatedOn", "Category"],
    "Space":     ["Name", "CreatedBy", "CreatedOn", "FloorName", "Category"],
    "Zone":      ["Name", "CreatedBy", "CreatedOn", "SpaceNames"],
    "Type":      ["Name", "CreatedBy", "CreatedOn", "Category", "AssetType"],
    "Component": ["Name", "CreatedBy", "CreatedOn", "TypeName", "Space"],
    "Document":  ["Name", "CreatedBy", "CreatedOn", "Stage", "SheetName", "RowName"],
    "Job":       ["Name", "CreatedBy", "CreatedOn", "TypeName", "Frequency"],
    "Resource":  ["Name", "CreatedBy", "CreatedOn"],
    "Spare":     ["Name", "CreatedBy", "CreatedOn", "TypeName"],
    "System":    ["Name", "CreatedBy", "CreatedOn", "ComponentNames"],
}

INTERNAL_FIELDS = ["ExternalSystem", "ExternalObject", "ExternalIdentifier"]


def _header_fill(sheet_name: str, col_name: str) -> PatternFill:
    required = REQUIRED_FIELDS.get(sheet_name, [])
    if col_name in required:
        colour = COLOURS["header_required"]
    elif col_name in INTERNAL_FIELDS:
        colour = COLOURS["header_internal"]
    else:
        colour = COLOURS["header_default"]
    return PatternFill("solid", fgColor=colour)


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = COLOURS["sheet_tab"]

    if df.empty:
        return

    cols = list(df.columns)

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill      = _header_fill(sheet_name, col)
        cell.font      = Font(bold=True, color="FFFFFF", size=9,
                              name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _thin_border()

    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        fill_colour = "FFFFFF" if ri % 2 == 0 else COLOURS["row_alt"]
        for ci, col in enumerate(cols, start=1):
            val  = row[col]
            sval = str(val) if pd.notna(val) else ""
            cell = ws.cell(row=ri, column=ci, value=sval)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = _thin_border()
            cell.font      = Font(size=9, name="Calibri")

            # Highlight TBC in required fields
            required = REQUIRED_FIELDS.get(sheet_name, [])
            if col in required and sval in ("TBC", ""):
                cell.fill = PatternFill("solid", fgColor=COLOURS["critical"])
            else:
                cell.fill = PatternFill("solid", fgColor=fill_colour)

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        max_len = max(
            len(str(col)),
            *[len(str(df.iloc[r][col])) for r in range(min(len(df), 50))]
        )
        width = min(max(max_len + 2, 12), 40)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── Freeze top row ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions


def export_xlsx(sheets: dict) -> bytes:
    """
    Export all COBie sheets to a formatted .xlsx file.
    Returns bytes suitable for st.download_button.
    """
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # COBie UK 2012 sheet order
    sheet_order = [
        "Contact", "Facility", "Floor", "Space", "Zone",
        "Type", "Component", "System", "Document",
        "Job", "Resource", "Spare",
    ]

    for name in sheet_order:
        if name in sheets and not sheets[name].empty:
            _write_sheet(wb, name, sheets[name])

    # Any extra sheets not in standard order
    for name, df in sheets.items():
        if name not in sheet_order and not df.empty:
            _write_sheet(wb, name, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
